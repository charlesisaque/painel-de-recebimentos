import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

ARQUIVO = "recebimentos.xlsx"

# cria ou abre a planilha
if not os.path.exists(ARQUIVO):
    wb = Workbook()
    ws = wb.active
    ws.append(["Data", "Hora", "Entregador", "Recebedor", "Empresa", "Código de Barras"])
    wb.save(ARQUIVO)

# carregar planilha
def carregar_dados():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    dados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dados.append(row)
    return dados

# salvar novo recebimento
def salvar_recebimento(entregador, recebedor, empresa, codigo):
    data = datetime.now().strftime("%d/%m/%Y")
    hora = datetime.now().strftime("%H:%M:%S")

    wb = load_workbook(ARQUIVO)
    ws = wb.active
    ws.append([data, hora, entregador, recebedor, empresa, codigo])
    wb.save(ARQUIVO)

# excluir recebimento
def excluir_recebimento(idx):
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    ws.delete_rows(idx+2)  # +2 pq primeira linha é cabeçalho
    wb.save(ARQUIVO)

# interface principal
def atualizar_tabela():
    for i in tree.get_children():
        tree.delete(i)
    for row in carregar_dados():
        tree.insert("", tk.END, values=row)

def novo_recebimento():
    def salvar():
        if not (entry_entregador.get() and entry_recebedor.get() and entry_empresa.get() and entry_codigo.get()):
            messagebox.showwarning("Atenção", "Preencha todos os campos!")
            return

        if messagebox.askyesno("Confirmação", "DESEJA CONFIRMAR NOVO RECEBIMENTO?\nNÃO ESQUEÇA DE CONFIRIR SE TODAS AS INFORMAÇÕES ESTÃO PREENCHIDAS CORRETAMENTE."):
            salvar_recebimento(entry_entregador.get(), entry_recebedor.get(), entry_empresa.get(), entry_codigo.get())
            atualizar_tabela()
            janela.destroy()

    janela = tk.Toplevel(root)
    janela.title("Novo Recebimento")

    tk.Label(janela, text="Entregador:").grid(row=0, column=0)
    entry_entregador = tk.Entry(janela, width=40)
    entry_entregador.grid(row=0, column=1)

    tk.Label(janela, text="Recebedor:").grid(row=1, column=0)
    entry_recebedor = tk.Entry(janela, width=40)
    entry_recebedor.grid(row=1, column=1)

    tk.Label(janela, text="Empresa:").grid(row=2, column=0)
    entry_empresa = tk.Entry(janela, width=40)
    entry_empresa.grid(row=2, column=1)

    tk.Label(janela, text="Código de Barras:").grid(row=3, column=0)
    entry_codigo = tk.Entry(janela, width=40)
    entry_codigo.grid(row=3, column=1)

    tk.Button(janela, text="Salvar", command=salvar).grid(row=4, column=0, columnspan=2, pady=10)

def deletar_recebimento():
    item = tree.selection()
    if not item:
        messagebox.showwarning("Atenção", "Selecione um recebimento para excluir!")
        return

    if messagebox.askyesno("Confirmação", "ATENÇÃO, DESEJA EXCLUIR ESSE RECEBIMENTO PERMANENTEMENTE?"):
        idx = tree.index(item[0])
        excluir_recebimento(idx)
        atualizar_tabela()

# janela principal
root = tk.Tk()
root.title("Controle de Encomendas")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

cols = ["Data", "Hora", "Entregador", "Recebedor", "Empresa", "Código de Barras"]
tree = ttk.Treeview(frame, columns=cols, show="headings")
for col in cols:
    tree.heading(col, text=col)
    tree.column(col, width=120)
tree.pack()

btn_frame = tk.Frame(root)
btn_frame.pack(pady=10)

tk.Button(btn_frame, text="Novo Recebimento", command=novo_recebimento).grid(row=0, column=0, padx=5)
tk.Button(btn_frame, text="Excluir Recebimento", command=deletar_recebimento).grid(row=0, column=1, padx=5)

atualizar_tabela()
root.mainloop()
